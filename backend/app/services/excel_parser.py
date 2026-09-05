"""Excel 解析与列映射（对应设计文档《Excel列映射方案.md》）

三层：表头同义词自动识别 → 单位/格式归一化 → 返回映射预览 + 数据。
"""
import csv
import io
import logging
import re
from typing import Any

logger = logging.getLogger(__name__)

# 表头同义词词典：标准字段 -> 表头关键词列表（依据 8 份真实债权表格扩充）
HEADER_SYNONYMS: dict[str, list[str]] = {
    "debtor_name": ["债务人", "债权项目", "借款人", "企业名称", "债务人名称", "债务企业", "借款企业", "主债务人", "借款人名称"],
    "principal_text": ["本金", "债权本金", "借款本金", "贷款本金", "本金余额", "收购本金", "本金万元", "债权本金（万元）"],
    "interest_text": ["利息", "债权利息", "欠息", "利息罚息", "应收利息", "利息余额", "利息(基准日）", "利息（0630）", "债权利息(截至"],
    "fees_text": ["费用", "诉讼费", "保全费", "实现债权费用"],
    "guaranty_type": ["担保类型", "担保方式", "担保", "担保形式", "担保情况"],
    "guarantor": ["保证人", "担保人", "连带保证", "保证人名称", "担保人（含抵押）", "保证人名称及保证金额"],
    "creditor": ["债权人", "原债权人", "出让方", "转让方", "原告", "出租方", "租赁公司"],
    "debt_type": ["债权类型", "业务类型", "产品类型", "资产类型", "租赁类型", "融资租赁"],
    "interest_method": ["计息方式", "利息计算方式", "利率", "利息算法", "租息方式", "计息条款"],
    "judgment_result": ["是否胜诉", "裁判结果", "判决结果", "胜诉情况", "诉讼结果"],
    "collateral": ["抵押物", "抵押物情况", "抵押物描述", "抵质押物", "担保物", "抵押情况", "抵/质押情况", "抵押物情况描述", "抵质押物情况", "抵押", "额外查封物"],
    "mortgagor": ["抵押人"],
    "collateral_type": ["抵押物类型", "抵质押资产分类", "抵押资产分类", "资产分类"],
    "judicial_status": ["执行法院", "司法状态", "诉讼状态", "法院", "受理法院", "管辖法院", "诉讼进度", "诉讼、查封情况"],
    "region": ["地区", "所在地", "借款人所处地级市", "地级市"],
    "batch": ["批次", "资产包名称", "资产包", "项目名称"],
    "loan_bank": ["贷款行", "贷款银行", "原贷款机构"],
    "listing_price_text": ["挂牌价", "起拍价", "评估价", "债权转让价", "转让价格"],
    "deadline": ["截止日期", "到期日", "拍卖时间", "报名截止", "挂牌截止", "转让基准日"],
    "interest_base_date": ["计息基准日", "转让基准日", "利息截止", "截至"],
    "mortgage_amount": ["抵押金额", "抵押价值", "最高额抵押", "贷款时估值"],
    "mortgage_rank": ["抵押顺位", "顺位", "第一顺位", "首押"],
    "seizure": ["查封情况", "查封", "首封", "轮候"],
    "collateral_status": ["抵押物现状", "现状", "已抵债", "已拍卖", "清场"],
    "debtor_status": ["债务人状态", "经营状态", "主债务人情况"],
    "guarantor_status": ["保证人情况", "担保人情况"],
    "extra_notes": ["备注", "说明", "其他", "附注", "债务人及抵押担保尽调情况"],
}

# 完整度判定：核心字段 + 重要字段
CORE_FIELDS = ["debtor_name", "principal_text", "collateral"]
IMPORTANT_FIELDS = ["interest_text", "guarantor", "judicial_status", "region", "collateral_type", "mortgagor"]

# 表头单位解析：列名里的单位词
_UNIT_PATTERN = re.compile(r"[（(]?\s*(万亿|亿|万|元)\s*[)）]?")

# 表头清洗：去空白、去括号内容、去单位词
def _clean_header(raw: str) -> str:
    t = re.sub(r"[（(][^）)]*[)）]", "", raw)  # 去括号内容
    t = _UNIT_PATTERN.sub("", t)
    return t.strip()


def build_mapping(headers: list[str]) -> dict[str, str]:
    """表头列表 -> {标准字段: 原始列名}，未识别的列省略（进 extra）"""
    mapping: dict[str, str] = {}
    for raw in headers:
        cleaned = _clean_header(raw)
        if not cleaned:
            continue
        for field, keywords in HEADER_SYNONYMS.items():
            if field in mapping:
                continue
            if any(kw in cleaned or cleaned in kw for kw in keywords):
                mapping[field] = raw
                break
    return mapping


def extract_interest_cutoff(header_raw: str | None) -> str | None:
    """从利息列表头提取计息截止日，如『债权利息(截至2025/4/20）』→ '2025-04-20'。

    权威来源（银行/AMC 表格、判决书）通常标注利息截止日，用于把利息续算到报告当日。
    返回 YYYY-MM-DD 或 None。
    """
    if not header_raw:
        return None
    m = re.search(r"(?:截至|截止|截止到|至)\s*(\d{4})[年/.-](\d{1,2})[月/.-](\d{1,2})日?", header_raw)
    if m:
        try:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return f"{y:04d}-{mo:02d}-{d:02d}"
        except ValueError:
            return None
    return None


def _detect_unit(raw: str) -> int:
    """表头里带 '万'/'亿' 返回"值→元"倍率，否则 1（值本身按元计）。

    与 _to_cents 内嵌单位倍率一致：万元=10^4（元），亿元=10^8，万亿=10^12。
    _to_cents 最终 ×100 转分。
    """
    m = re.search(r"[（(]\s*(万亿|亿|万|元)\s*[)）]", raw) or re.search(r"(万亿|亿|万)", raw)
    if not m:
        return 1
    return {"万亿": 10**12, "亿": 10**8, "万": 10**4, "元": 1}.get(m.group(1), 1)


def _to_cents(value: Any, unit_mult: int) -> int | None:
    """值转分。unit_mult 为表头单位倍率（"值→元"，万元=10^4）。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        num = float(value)
    else:
        t = str(value).strip().replace(",", "").replace("，", "")
        if not t or t in ("未知", "无", "面议", "-", "—", "待定"):
            return None
        # 去掉内嵌单位（如 "539万"）
        m = re.match(r"^([\d.]+)\s*(万亿|亿|万|元)?$", t)
        if not m:
            return None
        num = float(m.group(1))
        if m.group(2):
            unit_mult = {"万亿": 10**12, "亿": 10**8, "万": 10**4, "元": 1}.get(m.group(2), 1)
    return int(round(num * unit_mult * 100))


def _apply_unit_inheritance(rows: list[dict[str, Any]], unit_mult: dict[str, int]) -> None:
    """无单位金额列继承主金额列（principal_text）的单位倍率。

    行业常识（用户确认 2026-08-25）：同一张表/文档内计量单位一般一致。
    若本金列标了单位（如"债权本金（万元）"），而利息/费用列没标，
    则利息/费用按本金列同倍率解析（即同为万元），避免"749万 但利息 627元"的常识性错误。
    规则：取所有金额列中最大的"值→元"倍率（>1 才继承），无单位列按此重算。
    """
    amount_fields = ("principal_text", "interest_text", "fees_text", "listing_price_text")
    # 收集各列已识别的倍率，取最大（>1）作为同表默认单位
    table_mult = 1
    for f in amount_fields:
        m = unit_mult.get(f, 1)
        if m > table_mult:
            table_mult = m
    if table_mult <= 1:
        return  # 全表都无单位，无从继承
    for row in rows:
        main_val = row.get("principal_text")
        for field in amount_fields:
            if field in unit_mult and unit_mult[field] > 1:
                continue  # 该列自带单位，已正确解析
            value = row.get(field)
            # 纯数字金额（已按"元"解析）且明显小于本金（<10%，说明是"元"而非"万元"量级）
            # 此时按同表单位（万元）重算，避免"749万本金、利息627元"的常识性错误
            if isinstance(value, int) and value > 0 and isinstance(main_val, int) and main_val > 0:
                if value < main_val * 0.10:
                    row[field] = int(round(value * table_mult))
                    row[f"{field}_unit_inherited"] = True


def parse_excel(file_bytes: bytes, filename: str) -> tuple[list[dict[str, Any]], dict, list[str]]:
    """解析 Excel/Csv。

    Returns:
        (rows, mapping, unmapped_columns)
        rows: [{标准字段: 值}, ...]（金额已转分）
        mapping: {标准字段: 原始列名}
        unmapped_columns: 未识别的列名
    """
    if filename.lower().endswith(".csv"):
        return _parse_csv(file_bytes)
    return _parse_xlsx(file_bytes)


def _parse_xlsx(file_bytes: bytes) -> tuple[list[dict[str, Any]], dict, list[str]]:
    from openpyxl import load_workbook  # 延迟导入：无该依赖时纯函数仍可测试

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(h) if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return [], {}, []
    mapping = build_mapping(header)
    unmapped = [h for h in header if h and h not in mapping.values()]

    # 从利息列表头提取计息截止日（如 债权利息(截至2025/4/20）→ 2025-04-20）
    interest_cutoff = None
    interest_raw = mapping.get("interest_text")
    if interest_raw:
        interest_cutoff = extract_interest_cutoff(interest_raw)

    # 单位倍率
    unit_mult: dict[str, int] = {}
    for field, raw in mapping.items():
        if field in ("principal_text", "interest_text", "fees_text", "listing_price_text"):
            unit_mult[field] = _detect_unit(raw)

    rows = []
    for raw_row in rows_iter:
        if all(v is None or str(v).strip() == "" for v in raw_row):
            continue
        row: dict[str, Any] = {}
        for i, col_name in enumerate(header):
            if i >= len(raw_row):
                continue
            value = raw_row[i]
            field = next((f for f, r in mapping.items() if r == col_name), None)
            if not field:
                continue
            if field in unit_mult:
                row[field] = _to_cents(value, unit_mult[field])
            else:
                row[field] = str(value).strip() if value is not None else None
        if interest_cutoff and "interest_base_date" not in row:
            row["interest_base_date"] = interest_cutoff
        rows.append(row)
    _apply_unit_inheritance(rows, unit_mult)
    return rows, mapping, unmapped


def _parse_csv(file_bytes: bytes) -> tuple[list[dict[str, Any]], dict, list[str]]:
    # 编码探测：优先 UTF-8，失败 GBK
    text = None
    for enc in ("utf-8-sig", "gbk", "utf-8"):
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("无法识别 CSV 编码")

    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        return [], {}, []
    mapping = build_mapping(header)
    unmapped = [h for h in header if h and h not in mapping.values()]

    # 从利息列表头提取计息截止日
    interest_cutoff = None
    interest_raw = mapping.get("interest_text")
    if interest_raw:
        interest_cutoff = extract_interest_cutoff(interest_raw)

    unit_mult: dict[str, int] = {}
    for field, raw in mapping.items():
        if field in ("principal_text", "interest_text", "fees_text", "listing_price_text"):
            unit_mult[field] = _detect_unit(raw)

    rows = []
    for raw_row in reader:
        if all(not c.strip() for c in raw_row):
            continue
        row: dict[str, Any] = {}
        for i, col_name in enumerate(header):
            if i >= len(raw_row):
                continue
            value = raw_row[i].strip()
            if not value:
                continue
            field = next((f for f, r in mapping.items() if r == col_name), None)
            if not field:
                continue
            if field in unit_mult:
                row[field] = _to_cents(value, unit_mult[field])
            else:
                row[field] = value
        if interest_cutoff and "interest_base_date" not in row:
            row["interest_base_date"] = interest_cutoff
        rows.append(row)
    _apply_unit_inheritance(rows, unit_mult)
    return rows, mapping, unmapped
